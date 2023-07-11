import openpyxl
from openpyxl.styles import *
import os

path = "~/Desktop/Python-udemy/Employees.xlsx"
expanded_path = os.path.expanduser(
    path
)  # Expand the tilde to the user's home directory
workbook = openpyxl.load_workbook(expanded_path)
"""print(workbook.sheetnames)
print(workbook.active)"""
# workbook.create_sheet("test")
# workbook.save(expanded_path)
# sheet = workbook['test']
# del workbook['test']
# workbook.save(expanded_path)

sheet = workbook["EmployeeData"]
# print(sheet.active_cell)
# print(sheet.dimensions)
# print(sheet.sheet_format)
# print(sheet.sheet_properties)
# print(sheet.sheet_state)
# print(sheet.sheet_view)
"""print(sheet.max_column)
print(sheet.max_row)
for i in sheet.values:
    print(i)"""

# print(sheet['B7'].value)
"""print(sheet.cell(row=6, column=2).value)
cell = sheet['B2']
cell.value = 'David'
workbook.save(expanded_path)"""

# dir(openpyxl.styles) available styles

cell = sheet["B8"]

font = Font(color=Color(rgb="FF0000"), bold=True, italic=True)


cell.font = font

fill = PatternFill(fill_type="solid", bgColor="F7FE2E")

cell.fill = fill

border = Border(
    left=Side(border_style="double", color="322FEC"),
    right=Side(border_style="double", color="322FEC"),
    top=Side(border_style="double", color="322FEC"),
    bottom=Side(border_style="double", color="322FEC"),
)

cell.border = border

align = Alignment(horizontal="left")

cell.alignment = align

workbook.save(expanded_path)
